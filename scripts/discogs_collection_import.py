#!/usr/bin/env python3
"""
Discogs Collection Importer for VOD_Auctions.

Imports releases from a Discogs collection export (CSV or Excel) into Supabase.
Fetches full release data from Discogs API, matches against existing DB entries,
and provides a detailed preview before committing.

Supports two export formats:
  - CSV (standard Discogs export): Headers — Catalog#, Artist, Title, Label, Format, Rating, Released, release_id, ...
  - Excel (custom export, no headers): Artist, Title, CatalogNumber, Label, Format, Condition, Year, DiscogsID

Usage:
  # Simulate (default) — fetch + match + report, no DB write
  python3 discogs_collection_import.py --file ../discogs/export.xlsx --collection "Sammlung Müller"

  # Commit to DB
  python3 discogs_collection_import.py --file ../discogs/export.xlsx --collection "Sammlung Müller" --commit

  # Test with limit + markdown report
  python3 discogs_collection_import.py --file ../discogs/export.xlsx --collection "Test" --limit 20 --report ../discogs/report.md

  # Re-run from cache (skip API fetch)
  python3 discogs_collection_import.py --file ../discogs/export.xlsx --collection "Test" --skip-fetch
"""

import argparse
import csv
import json
import os
import signal
import sys
import time
import uuid
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import requests

# Add parent to path for shared imports
sys.path.insert(0, str(Path(__file__).parent))
from shared import (
    DISCOGS_BASE,
    DISCOGS_HEADERS,
    RateLimiter,
    slugify,
    decode_entities,
)

# psycopg2 and DB functions are imported lazily (only needed for --commit)
def _get_db_deps():
    """Lazy import of psycopg2 and DB helpers (avoids import errors when only simulating)."""
    import psycopg2
    import psycopg2.extras
    from shared import get_pg_connection, log_sync
    return psycopg2, get_pg_connection, log_sync

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

DATA_DIR = Path(__file__).parent / "data"
CACHE_FILE = DATA_DIR / "discogs_import_cache.json"
PROGRESS_FILE = DATA_DIR / "discogs_import_progress.json"

DB_DISCOGS_IDS_FILE = DATA_DIR / "db_discogs_ids.json"
DB_UNLINKED_FILE = DATA_DIR / "db_unlinked_releases.json"

DISCOGS_CONDITION_MAP = {
    1: "Poor (P)",
    2: "Fair (F)",
    3: "Good (G)",
    4: "Very Good (VG)",
    5: "Near Mint (NM) or Mint (M)",
}

# Map Discogs format names to our format_group enum
FORMAT_GROUP_MAP = {
    "Vinyl": "LP",
    "CD": "CD",
    "Cassette": "CASSETTE",
    "DVD": "VHS",
    "Blu-ray": "VHS",
    "Box Set": "BOXSET",
    "File": "DIGITAL",
    "Reel-To-Reel": "REEL",
}

# Graceful shutdown
_shutdown = False


def _handle_signal(sig, frame):
    global _shutdown
    print("\n[!] Shutdown requested — saving progress...")
    _shutdown = True


signal.signal(signal.SIGINT, _handle_signal)
signal.signal(signal.SIGTERM, _handle_signal)


# ---------------------------------------------------------------------------
# Excel / CSV Parsing
# ---------------------------------------------------------------------------

def parse_export_file(filepath: str) -> list[dict]:
    """Parse a Discogs export file (CSV or Excel) into a list of dicts."""
    path = Path(filepath)
    if not path.exists():
        print(f"ERROR: File not found: {filepath}")
        sys.exit(1)

    ext = path.suffix.lower()
    if ext == ".csv":
        return _parse_csv(path)
    elif ext in (".xlsx", ".xls"):
        return _parse_excel(path)
    else:
        print(f"ERROR: Unsupported file format: {ext} (expected .csv or .xlsx)")
        sys.exit(1)


def _parse_csv(path: Path) -> list[dict]:
    """Parse standard Discogs CSV export (with headers)."""
    rows = []
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            discogs_id = row.get("release_id", "").strip()
            if not discogs_id or not discogs_id.isdigit():
                continue
            rating = row.get("Rating", "").strip()
            rows.append({
                "artist": row.get("Artist", "").strip(),
                "title": row.get("Title", "").strip(),
                "catalog_number": row.get("Catalog#", "").strip(),
                "label": row.get("Label", "").strip(),
                "format": row.get("Format", "").strip(),
                "condition": int(rating) if rating.isdigit() else None,
                "year": _parse_year(row.get("Released", "")),
                "discogs_id": int(discogs_id),
                "collection_folder": row.get("CollectionFolder", "").strip(),
                "date_added": row.get("Date Added", "").strip(),
            })
    return rows


def _parse_excel(path: Path) -> list[dict]:
    """Parse Excel export (no headers, fixed column order)."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(values_only=True):
        if len(row) < 8:
            continue
        discogs_id = row[7]
        if discogs_id is None:
            continue
        try:
            discogs_id = int(discogs_id)
        except (ValueError, TypeError):
            continue
        condition = row[5]
        rows.append({
            "artist": str(row[0] or "").strip(),
            "title": str(row[1] or "").strip(),
            "catalog_number": str(row[2] or "").strip(),
            "label": str(row[3] or "").strip(),
            "format": str(row[4] or "").strip(),
            "condition": int(condition) if condition is not None else None,
            "year": _parse_year(row[6]),
            "discogs_id": discogs_id,
        })
    wb.close()
    return rows


def _parse_year(val) -> int | None:
    if val is None:
        return None
    try:
        y = int(str(val).strip())
        return y if 1900 <= y <= 2100 else None
    except (ValueError, TypeError):
        return None


def deduplicate(rows: list[dict]) -> list[dict]:
    """Deduplicate by discogs_id, keeping the entry with highest condition rating."""
    seen: dict[int, dict] = {}
    for row in rows:
        did = row["discogs_id"]
        if did not in seen:
            seen[did] = row
        else:
            existing_cond = seen[did].get("condition") or 0
            new_cond = row.get("condition") or 0
            if new_cond > existing_cond:
                seen[did] = row
    return list(seen.values())


# ---------------------------------------------------------------------------
# Discogs API Fetch
# ---------------------------------------------------------------------------

def fetch_discogs_data(
    rows: list[dict],
    cache: dict,
    rate_limiter: RateLimiter,
    limit: int | None = None,
) -> dict:
    """Fetch full release data from Discogs API. Returns updated cache dict."""
    to_fetch = [r for r in rows if str(r["discogs_id"]) not in cache]
    if limit is not None:
        to_fetch = to_fetch[:limit]

    if not to_fetch:
        print(f"  All {len(rows)} releases already cached.")
        return cache

    total = len(to_fetch)
    print(f"  Fetching {total} releases from Discogs API (~{total // 40 + 1} min)...")

    errors = []
    for i, row in enumerate(to_fetch):
        if _shutdown:
            print(f"  Stopped at {i}/{total} — progress saved.")
            break

        did = row["discogs_id"]
        rate_limiter.wait()

        try:
            resp = requests.get(
                f"{DISCOGS_BASE}/releases/{did}",
                headers=DISCOGS_HEADERS,
                timeout=30,
            )

            if resp.status_code == 429:
                retry_after = int(resp.headers.get("Retry-After", 60))
                print(f"  Rate limited — waiting {retry_after}s...")
                time.sleep(retry_after)
                rate_limiter.wait()
                resp = requests.get(
                    f"{DISCOGS_BASE}/releases/{did}",
                    headers=DISCOGS_HEADERS,
                    timeout=30,
                )

            if resp.status_code == 200:
                data = resp.json()
                cache[str(did)] = {
                    "title": data.get("title", ""),
                    "year": data.get("year", 0),
                    "country": data.get("country", ""),
                    "artists": [
                        {"name": a.get("name", ""), "id": a.get("id")}
                        for a in data.get("artists", [])
                    ],
                    "extraartists": [
                        {"name": a.get("name", ""), "id": a.get("id"), "role": a.get("role", "")}
                        for a in data.get("extraartists", [])
                    ],
                    "labels": [
                        {"name": l.get("name", ""), "catno": l.get("catno", ""), "id": l.get("id")}
                        for l in data.get("labels", [])
                    ],
                    "formats": [
                        {
                            "name": f.get("name", ""),
                            "descriptions": f.get("descriptions", []),
                            "qty": f.get("qty", "1"),
                        }
                        for f in data.get("formats", [])
                    ],
                    "tracklist": [
                        {"position": t.get("position", ""), "title": t.get("title", ""), "duration": t.get("duration", "")}
                        for t in data.get("tracklist", [])
                    ],
                    "genres": data.get("genres", []),
                    "styles": data.get("styles", []),
                    "community": {
                        "have": data.get("community", {}).get("have", 0),
                        "want": data.get("community", {}).get("want", 0),
                    },
                    "lowest_price": data.get("lowest_price"),
                    "num_for_sale": data.get("num_for_sale", 0),
                    "images": [
                        {"uri": img.get("uri", ""), "type": img.get("type", "")}
                        for img in data.get("images", [])[:10]  # max 10 images
                    ],
                    "notes": data.get("notes", ""),
                    "data_quality": data.get("data_quality", ""),
                    "fetched_at": datetime.now(timezone.utc).isoformat(),
                }
            elif resp.status_code == 404:
                cache[str(did)] = {"error": "not_found", "fetched_at": datetime.now(timezone.utc).isoformat()}
                errors.append((did, "404 Not Found"))
            else:
                errors.append((did, f"HTTP {resp.status_code}"))

        except requests.RequestException as e:
            errors.append((did, str(e)))

        if (i + 1) % 50 == 0 or i + 1 == total:
            print(f"  [{i + 1}/{total}] fetched...")
            _save_cache(cache)

    _save_cache(cache)

    if errors:
        print(f"  {len(errors)} errors during fetch:")
        for did, err in errors[:10]:
            print(f"    discogs:{did} — {err}")
        if len(errors) > 10:
            print(f"    ... and {len(errors) - 10} more")

    return cache


def _save_cache(cache: dict):
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with open(CACHE_FILE, "w") as f:
        json.dump(cache, f, indent=2, ensure_ascii=False)


def _load_cache() -> dict:
    if CACHE_FILE.exists():
        with open(CACHE_FILE) as f:
            return json.load(f)
    return {}


# ---------------------------------------------------------------------------
# DB Matching
# ---------------------------------------------------------------------------

def match_against_db(rows: list[dict], cache: dict) -> dict:
    """Match import rows against existing DB entries using local snapshot files.

    Requires pre-exported snapshots in data/:
      - db_discogs_ids.json: {discogs_id: release_id, ...}
      - db_unlinked_releases.json: [{id, title, catalog_number, artist_name}, ...]

    Regenerate snapshots via Supabase MCP or direct DB query when DB changes significantly.
    """
    print("\n── Phase 2: Matching against database ──")

    # 1. Load existing discogs_id -> release_id mapping
    if not DB_DISCOGS_IDS_FILE.exists():
        print(f"  ERROR: {DB_DISCOGS_IDS_FILE} not found.")
        print(f"  Export it first: SELECT id, discogs_id FROM \"Release\" WHERE discogs_id IS NOT NULL")
        sys.exit(1)

    with open(DB_DISCOGS_IDS_FILE) as f:
        existing_by_discogs_str = json.load(f)  # {"665510": "legacy-release-123", ...}
    # Convert string keys to int for matching
    existing_by_discogs = {int(k): v for k, v in existing_by_discogs_str.items()}
    print(f"  Loaded {len(existing_by_discogs)} releases with discogs_id (from snapshot)")

    # 2. Load unlinked releases for fuzzy matching
    if DB_UNLINKED_FILE.exists():
        with open(DB_UNLINKED_FILE) as f:
            unlinked = json.load(f)
        print(f"  Loaded {len(unlinked)} releases without discogs_id (for fuzzy match)")
    else:
        unlinked = []
        print(f"  WARNING: {DB_UNLINKED_FILE} not found — fuzzy matching disabled")

    # Build fuzzy lookup index: normalized (artist, title, catno) -> release
    fuzzy_index = {}
    for rel in unlinked:
        key = _fuzzy_key(rel.get("artist_name", ""), rel.get("title", ""), rel.get("catalog_number", ""))
        if key:
            fuzzy_index[key] = rel

    results = {"existing": [], "linkable": [], "new": [], "skipped": []}

    for row in rows:
        did = row["discogs_id"]
        api_data = cache.get(str(did), {})

        if api_data.get("error"):
            results["skipped"].append({**row, "skip_reason": api_data["error"]})
            continue

        # Match 1: Exact discogs_id
        if did in existing_by_discogs:
            results["existing"].append({
                **row,
                "db_release_id": existing_by_discogs[did],
                "api_data": api_data,
            })
            continue

        # Match 2: Fuzzy by artist+title+catno
        key = _fuzzy_key(row["artist"], row["title"], row["catalog_number"])
        if key and key in fuzzy_index:
            db_rel = fuzzy_index[key]
            results["linkable"].append({
                **row,
                "db_release_id": db_rel["id"],
                "db_title": db_rel["title"],
                "match_key": key,
                "api_data": api_data,
            })
            continue

        # Match 3: No match → new
        results["new"].append({**row, "api_data": api_data})
    return results


def _fuzzy_key(artist: str, title: str, catno: str) -> str | None:
    """Create a normalized lookup key for fuzzy matching."""
    a = _normalize(artist)
    t = _normalize(title)
    c = _normalize(catno)
    if not a or not t:
        return None
    return f"{a}|{t}|{c}"


def _normalize(s: str) -> str:
    if not s:
        return ""
    s = decode_entities(s).lower().strip()
    s = s.replace("'", "").replace('"', "").replace(".", "").replace(",", "")
    return " ".join(s.split())


# ---------------------------------------------------------------------------
# Report Generation
# ---------------------------------------------------------------------------

def print_report(results: dict, source_file: str, collection: str):
    """Print formatted preview report to terminal."""
    total = sum(len(v) for v in results.values())
    print()
    print("═" * 60)
    print(" DISCOGS COLLECTION IMPORT — PREVIEW")
    print("═" * 60)
    print(f"  Source:       {Path(source_file).name}")
    print(f"  Collection:   {collection}")
    print(f"  Date:         {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print()
    print(f"  Total unique:           {total:>6}")
    print(f"  ─────────────────────────────")
    print(f"  EXISTING (in DB):       {len(results['existing']):>6}  ← price/community update")
    print(f"  LINKABLE (match found): {len(results['linkable']):>6}  ← add discogs_id + prices")
    print(f"  NEW (to import):        {len(results['new']):>6}  ← full import")
    print(f"  SKIPPED (API error):    {len(results['skipped']):>6}")
    print(f"  ─────────────────────────────")
    print()

    if results["new"]:
        print("  NEW RELEASES (first 30):")
        for i, r in enumerate(results["new"][:30], 1):
            year = r.get("year") or "?"
            fmt = r.get("format", "?")
            print(f"    {i:>3}. {r['artist']} — {r['title']} ({year}, {fmt}) [discogs:{r['discogs_id']}]")
        if len(results["new"]) > 30:
            print(f"    ... and {len(results['new']) - 30} more")
        print()

    if results["linkable"]:
        print("  LINKABLE (existing release ↔ Discogs ID):")
        for i, r in enumerate(results["linkable"][:20], 1):
            print(f"    {i:>3}. {r['db_release_id']} ↔ discogs:{r['discogs_id']} ({r['artist']} — {r['title']})")
        if len(results["linkable"]) > 20:
            print(f"    ... and {len(results['linkable']) - 20} more")
        print()

    if results["skipped"]:
        print("  SKIPPED:")
        for r in results["skipped"][:10]:
            print(f"    discogs:{r['discogs_id']} — {r.get('skip_reason', '?')} ({r['artist']} — {r['title']})")
        print()


def generate_markdown_report(results: dict, source_file: str, collection: str, output_path: str):
    """Generate a detailed .md report file."""
    total = sum(len(v) for v in results.values())
    lines = [
        f"# Discogs Collection Import — Report",
        f"",
        f"| Field | Value |",
        f"|-------|-------|",
        f"| Source | `{Path(source_file).name}` |",
        f"| Collection | {collection} |",
        f"| Date | {datetime.now().strftime('%Y-%m-%d %H:%M')} |",
        f"| Total unique | {total} |",
        f"| Mode | **Simulation** (no DB write) |",
        f"",
        f"## Summary",
        f"",
        f"| Category | Count | Action |",
        f"|----------|------:|--------|",
        f"| EXISTING | {len(results['existing'])} | Update prices + community data |",
        f"| LINKABLE | {len(results['linkable'])} | Add discogs_id + prices to existing release |",
        f"| NEW | {len(results['new'])} | Full import (Release + Artist + Label + Tracks) |",
        f"| SKIPPED | {len(results['skipped'])} | API error — not importable |",
        f"| **Total** | **{total}** | |",
        f"",
    ]

    # NEW section
    if results["new"]:
        lines.append(f"## NEW Releases ({len(results['new'])})")
        lines.append("")
        lines.append("| # | Artist | Title | Year | Format | Discogs ID |")
        lines.append("|--:|--------|-------|------|--------|------------|")
        for i, r in enumerate(results["new"], 1):
            year = r.get("year") or "?"
            fmt = r.get("format", "?")
            lines.append(f"| {i} | {r['artist']} | {r['title']} | {year} | {fmt} | {r['discogs_id']} |")
        lines.append("")

    # LINKABLE section
    if results["linkable"]:
        lines.append(f"## LINKABLE Releases ({len(results['linkable'])})")
        lines.append("")
        lines.append("| # | DB Release ID | Artist | Title | Discogs ID |")
        lines.append("|--:|---------------|--------|-------|------------|")
        for i, r in enumerate(results["linkable"], 1):
            lines.append(f"| {i} | `{r['db_release_id']}` | {r['artist']} | {r['title']} | {r['discogs_id']} |")
        lines.append("")

    # EXISTING section (summary only)
    if results["existing"]:
        lines.append(f"## EXISTING Releases ({len(results['existing'])})")
        lines.append("")
        lines.append(f"These {len(results['existing'])} releases already have a matching `discogs_id` in the database.")
        lines.append("On commit, only prices and community data would be updated.")
        lines.append("")
        lines.append("<details>")
        lines.append("<summary>Show all existing matches</summary>")
        lines.append("")
        lines.append("| # | DB Release ID | Artist | Title | Discogs ID |")
        lines.append("|--:|---------------|--------|-------|------------|")
        for i, r in enumerate(results["existing"], 1):
            lines.append(f"| {i} | `{r['db_release_id']}` | {r['artist']} | {r['title']} | {r['discogs_id']} |")
        lines.append("")
        lines.append("</details>")
        lines.append("")

    # SKIPPED section
    if results["skipped"]:
        lines.append(f"## SKIPPED ({len(results['skipped'])})")
        lines.append("")
        lines.append("| # | Artist | Title | Discogs ID | Reason |")
        lines.append("|--:|--------|-------|------------|--------|")
        for i, r in enumerate(results["skipped"], 1):
            lines.append(f"| {i} | {r['artist']} | {r['title']} | {r['discogs_id']} | {r.get('skip_reason', '?')} |")
        lines.append("")

    # Sample API data for new releases
    if results["new"]:
        lines.append("## Sample API Data (first 3 NEW releases)")
        lines.append("")
        for r in results["new"][:3]:
            api = r.get("api_data", {})
            if not api:
                continue
            lines.append(f"### {r['artist']} — {r['title']} (discogs:{r['discogs_id']})")
            lines.append("")
            lines.append(f"- **Country:** {api.get('country', '?')}")
            lines.append(f"- **Year:** {api.get('year', '?')}")
            lines.append(f"- **Genres:** {', '.join(api.get('genres', []))}")
            lines.append(f"- **Styles:** {', '.join(api.get('styles', []))}")
            lines.append(f"- **Formats:** {', '.join(f.get('name', '') for f in api.get('formats', []))}")
            lines.append(f"- **Community:** {api.get('community', {}).get('have', 0)} have / {api.get('community', {}).get('want', 0)} want")
            lines.append(f"- **Lowest Price:** {api.get('lowest_price', 'N/A')}")
            lines.append(f"- **For Sale:** {api.get('num_for_sale', 0)}")
            tracks = api.get("tracklist", [])
            if tracks:
                lines.append(f"- **Tracklist:** {len(tracks)} tracks")
                for t in tracks[:5]:
                    lines.append(f"  - {t.get('position', '')} {t.get('title', '')} ({t.get('duration', '')})")
                if len(tracks) > 5:
                    lines.append(f"  - ... +{len(tracks) - 5} more")
            images = api.get("images", [])
            lines.append(f"- **Images:** {len(images)}")
            lines.append("")

    report = "\n".join(lines)
    Path(output_path).write_text(report, encoding="utf-8")
    print(f"  Report saved to: {output_path}")


# ---------------------------------------------------------------------------
# DB Import (--commit)
# ---------------------------------------------------------------------------

def _map_discogs_format(api_data: dict) -> str:
    """Map Discogs format to our format_group enum."""
    formats = api_data.get("formats", [])
    if not formats:
        return "OTHER"
    name = formats[0].get("name", "")
    return FORMAT_GROUP_MAP.get(name, "OTHER")


def _ensure_import_log_table(conn):
    """Create import_log table if it doesn't exist."""
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS import_log (
            id TEXT PRIMARY KEY,
            import_type TEXT NOT NULL,
            collection_name TEXT,
            import_source TEXT NOT NULL,
            run_id TEXT NOT NULL,
            release_id TEXT,
            discogs_id INTEGER,
            action TEXT NOT NULL,
            data_snapshot JSONB,
            created_at TIMESTAMPTZ DEFAULT NOW()
        )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_import_log_run ON import_log(run_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_import_log_release ON import_log(release_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_import_log_collection ON import_log(collection_name)")
    conn.commit()


def commit_import(results: dict, collection: str, source_file: str):
    """Write matched results to DB."""
    psycopg2, get_pg_connection, log_sync = _get_db_deps()
    conn = get_pg_connection()
    cur = conn.cursor()
    run_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()

    _ensure_import_log_table(conn)

    print(f"\n── Phase 3: Committing to DB (run_id: {run_id[:8]}...) ──")

    counters = {"inserted": 0, "linked": 0, "updated": 0, "skipped": 0, "errors": 0}

    # --- NEW releases ---
    for row in results["new"]:
        if _shutdown:
            break
        try:
            api = row.get("api_data", {})
            if not api or api.get("error"):
                counters["skipped"] += 1
                continue

            did = row["discogs_id"]
            release_id = f"discogs-release-{did}"

            # Ensure Artist
            artist_id = _ensure_artist(cur, api, row)

            # Ensure Label
            label_id = _ensure_label(cur, api, row)

            # Insert Release
            format_group = _map_discogs_format(api)
            cur.execute("""
                INSERT INTO "Release" (
                    id, title, slug, "artistId", "labelId",
                    "catalogNumber", year, country, format_group,
                    discogs_id, discogs_lowest_price, discogs_num_for_sale,
                    discogs_have, discogs_want, discogs_last_synced,
                    product_category, "createdAt", "updatedAt"
                ) VALUES (
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    'release', NOW(), NOW()
                )
                ON CONFLICT (id) DO NOTHING
            """, (
                release_id,
                api.get("title") or row["title"],
                slugify(f"{row['artist']} {row['title']}"),
                artist_id,
                label_id,
                row["catalog_number"],
                api.get("year") or row.get("year"),
                api.get("country", ""),
                format_group,
                did,
                api.get("lowest_price"),
                api.get("num_for_sale", 0),
                api.get("community", {}).get("have", 0),
                api.get("community", {}).get("want", 0),
                now,
            ))

            # ReleaseArtist junction
            for artist_data in api.get("artists", []):
                a_id = _ensure_artist_by_discogs(cur, artist_data)
                if a_id:
                    cur.execute("""
                        INSERT INTO "ReleaseArtist" (id, "releaseId", "artistId", role, "createdAt")
                        VALUES (%s, %s, %s, 'performer', NOW())
                        ON CONFLICT ("releaseId", "artistId") DO NOTHING
                    """, (f"dra-{did}-{artist_data.get('id', 0)}", release_id, a_id))

            # Tracklist
            for track in api.get("tracklist", []):
                t_title = track.get("title", "")
                if not t_title:
                    continue
                cur.execute("""
                    INSERT INTO "Track" (id, "releaseId", position, title, duration, "createdAt")
                    VALUES (%s, %s, %s, %s, %s, NOW())
                    ON CONFLICT DO NOTHING
                """, (
                    f"dt-{did}-{track.get('position', '0')}",
                    release_id,
                    track.get("position", ""),
                    t_title,
                    track.get("duration", ""),
                ))

            # Import log
            _log_import(cur, run_id, collection, source_file, release_id, did, "inserted", row, api)
            counters["inserted"] += 1

        except Exception as e:
            print(f"  ERROR inserting discogs:{row['discogs_id']}: {e}")
            counters["errors"] += 1
            conn.rollback()
            continue

        if counters["inserted"] % 50 == 0:
            conn.commit()
            print(f"  Inserted: {counters['inserted']}...")

    # --- LINKABLE releases ---
    for row in results["linkable"]:
        if _shutdown:
            break
        try:
            api = row.get("api_data", {})
            did = row["discogs_id"]
            db_id = row["db_release_id"]

            cur.execute("""
                UPDATE "Release" SET
                    discogs_id = %s,
                    discogs_lowest_price = %s,
                    discogs_num_for_sale = %s,
                    discogs_have = %s,
                    discogs_want = %s,
                    discogs_last_synced = %s,
                    "updatedAt" = NOW()
                WHERE id = %s
            """, (
                did,
                api.get("lowest_price"),
                api.get("num_for_sale", 0),
                api.get("community", {}).get("have", 0),
                api.get("community", {}).get("want", 0),
                now,
                db_id,
            ))
            _log_import(cur, run_id, collection, source_file, db_id, did, "linked", row, api)
            counters["linked"] += 1
        except Exception as e:
            print(f"  ERROR linking discogs:{row['discogs_id']}: {e}")
            counters["errors"] += 1

    # --- EXISTING releases (price update) ---
    for row in results["existing"]:
        if _shutdown:
            break
        try:
            api = row.get("api_data", {})
            did = row["discogs_id"]
            db_id = row["db_release_id"]

            cur.execute("""
                UPDATE "Release" SET
                    discogs_lowest_price = %s,
                    discogs_num_for_sale = %s,
                    discogs_have = %s,
                    discogs_want = %s,
                    discogs_last_synced = %s,
                    "updatedAt" = NOW()
                WHERE id = %s
            """, (
                api.get("lowest_price"),
                api.get("num_for_sale", 0),
                api.get("community", {}).get("have", 0),
                api.get("community", {}).get("want", 0),
                now,
                db_id,
            ))
            _log_import(cur, run_id, collection, source_file, db_id, did, "updated", row, api)
            counters["updated"] += 1
        except Exception as e:
            print(f"  ERROR updating discogs:{row['discogs_id']}: {e}")
            counters["errors"] += 1

    conn.commit()

    # Summary sync_log entry
    log_sync(conn, None, "discogs_collection_import", {
        "run_id": run_id,
        "collection": collection,
        "source_file": Path(source_file).name,
        "inserted": counters["inserted"],
        "linked": counters["linked"],
        "updated": counters["updated"],
        "skipped": counters["skipped"],
        "errors": counters["errors"],
    })

    conn.close()

    print(f"\n  Done!")
    print(f"  Inserted: {counters['inserted']}")
    print(f"  Linked:   {counters['linked']}")
    print(f"  Updated:  {counters['updated']}")
    print(f"  Skipped:  {counters['skipped']}")
    print(f"  Errors:   {counters['errors']}")
    print(f"  Run ID:   {run_id}")


def _ensure_artist(cur, api_data: dict, row: dict) -> str | None:
    """Ensure primary artist exists, return artist ID."""
    artists = api_data.get("artists", [])
    if artists:
        return _ensure_artist_by_discogs(cur, artists[0])
    # Fallback: use Excel artist name
    name = row.get("artist", "").strip()
    if not name:
        return None
    slug = slugify(name)
    cur.execute('SELECT id FROM "Artist" WHERE slug = %s', (slug,))
    existing = cur.fetchone()
    if existing:
        return existing[0]
    artist_id = f"import-artist-{slug}"
    cur.execute("""
        INSERT INTO "Artist" (id, name, slug, "createdAt", "updatedAt")
        VALUES (%s, %s, %s, NOW(), NOW())
        ON CONFLICT (id) DO NOTHING
        RETURNING id
    """, (artist_id, name, slug))
    result = cur.fetchone()
    return result[0] if result else artist_id


def _ensure_artist_by_discogs(cur, artist_data: dict) -> str | None:
    """Ensure artist exists by Discogs artist data."""
    discogs_aid = artist_data.get("id")
    name = artist_data.get("name", "").strip()
    if not name:
        return None

    # Clean Discogs artist name suffixes like " (2)"
    import re
    name = re.sub(r"\s*\(\d+\)$", "", name)

    if discogs_aid:
        artist_id = f"discogs-artist-{discogs_aid}"
        cur.execute('SELECT id FROM "Artist" WHERE id = %s', (artist_id,))
        if cur.fetchone():
            return artist_id

    # Check by slug
    slug = slugify(name)
    cur.execute('SELECT id FROM "Artist" WHERE slug = %s', (slug,))
    existing = cur.fetchone()
    if existing:
        return existing[0]

    # Insert new
    artist_id = f"discogs-artist-{discogs_aid}" if discogs_aid else f"import-artist-{slug}"
    cur.execute("""
        INSERT INTO "Artist" (id, name, slug, "createdAt", "updatedAt")
        VALUES (%s, %s, %s, NOW(), NOW())
        ON CONFLICT (id) DO NOTHING
        RETURNING id
    """, (artist_id, name, slug))
    result = cur.fetchone()
    return result[0] if result else artist_id


def _ensure_label(cur, api_data: dict, row: dict) -> str | None:
    """Ensure primary label exists, return label ID."""
    labels = api_data.get("labels", [])
    if labels:
        label = labels[0]
        discogs_lid = label.get("id")
        name = label.get("name", "").strip()
        if not name:
            return None

        # Clean suffixes like " (2)"
        import re
        name = re.sub(r"\s*\(\d+\)$", "", name)

        if discogs_lid:
            label_id = f"discogs-label-{discogs_lid}"
            cur.execute('SELECT id FROM "Label" WHERE id = %s', (label_id,))
            if cur.fetchone():
                return label_id

        slug = slugify(name)
        cur.execute('SELECT id FROM "Label" WHERE slug = %s', (slug,))
        existing = cur.fetchone()
        if existing:
            return existing[0]

        label_id = f"discogs-label-{discogs_lid}" if discogs_lid else f"import-label-{slug}"
        cur.execute("""
            INSERT INTO "Label" (id, name, slug, "createdAt", "updatedAt")
            VALUES (%s, %s, %s, NOW(), NOW())
            ON CONFLICT (id) DO NOTHING
            RETURNING id
        """, (label_id, name, slug))
        result = cur.fetchone()
        return result[0] if result else label_id

    # Fallback from Excel
    name = row.get("label", "").strip()
    if not name:
        return None
    slug = slugify(name)
    cur.execute('SELECT id FROM "Label" WHERE slug = %s', (slug,))
    existing = cur.fetchone()
    if existing:
        return existing[0]
    return None


def _log_import(cur, run_id, collection, source_file, release_id, discogs_id, action, row, api_data):
    """Write a single import_log entry."""
    log_id = f"ilog-{run_id[:8]}-{discogs_id}"
    snapshot = {
        "excel": {k: v for k, v in row.items() if k != "api_data"},
        "api_summary": {
            "title": api_data.get("title", ""),
            "year": api_data.get("year"),
            "country": api_data.get("country", ""),
            "genres": api_data.get("genres", []),
            "community": api_data.get("community", {}),
            "lowest_price": api_data.get("lowest_price"),
        } if api_data else None,
    }
    cur.execute("""
        INSERT INTO import_log (id, import_type, collection_name, import_source, run_id, release_id, discogs_id, action, data_snapshot)
        VALUES (%s, 'discogs_collection', %s, %s, %s, %s, %s, %s, %s::jsonb)
        ON CONFLICT (id) DO NOTHING
    """, (log_id, collection, Path(source_file).name, run_id, release_id, discogs_id, action, json.dumps(snapshot, ensure_ascii=False)))


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Import Discogs collection export into VOD Supabase DB")
    parser.add_argument("--file", required=True, help="Path to Discogs export file (.csv or .xlsx)")
    parser.add_argument("--collection", required=True, help="Collection name for tracking (e.g. 'Sammlung Müller')")
    parser.add_argument("--commit", action="store_true", help="Actually write to DB (default: simulate only)")
    parser.add_argument("--skip-fetch", action="store_true", help="Skip API fetch, use cached data only")
    parser.add_argument("--limit", type=int, default=None, help="Limit number of entries to process")
    parser.add_argument("--report", type=str, default=None, help="Save markdown report to this path")
    args = parser.parse_args()

    print("═" * 60)
    print(" DISCOGS COLLECTION IMPORTER")
    print("═" * 60)
    print(f"  File:       {args.file}")
    print(f"  Collection: {args.collection}")
    print(f"  Mode:       {'COMMIT' if args.commit else 'SIMULATE'}")
    if args.limit:
        print(f"  Limit:      {args.limit}")
    print()

    # Phase 1: Parse Excel/CSV
    print("── Phase 1: Parsing export file ──")
    rows = parse_export_file(args.file)
    print(f"  Parsed {len(rows)} entries")

    rows = deduplicate(rows)
    print(f"  {len(rows)} unique Discogs IDs after deduplication")

    if args.limit:
        rows = rows[:args.limit]
        print(f"  Limited to {len(rows)} entries")

    # Phase 1b: Fetch from Discogs API
    cache = _load_cache()
    if not args.skip_fetch:
        print("\n── Phase 1b: Fetching from Discogs API ──")
        rate_limiter = RateLimiter(max_calls=40, period=60)
        cache = fetch_discogs_data(rows, cache, rate_limiter)
    else:
        print(f"\n  Skipping API fetch — using {len(cache)} cached entries")

    if _shutdown:
        print("\nShutdown — exiting before match phase.")
        return

    # Phase 2: Match against DB
    results = match_against_db(rows, cache)

    # Print report
    print_report(results, args.file, args.collection)

    # Save markdown report if requested
    if args.report:
        generate_markdown_report(results, args.file, args.collection, args.report)

    # Phase 3: Commit or simulate
    if args.commit:
        answer = input("  Proceed with import? [y/N] ").strip().lower()
        if answer == "y":
            commit_import(results, args.collection, args.file)
        else:
            print("  Import cancelled.")
    else:
        print("  Simulation complete. Use --commit to write to DB.")


if __name__ == "__main__":
    main()
